from pyrogram import Client, filters
from pyrogram.types import InlineKeyboardMarkup, InlineKeyboardButton
from pyrogram.enums import ParseMode
from openpyxl import load_workbook
import unicodedata
import os
import random

api_id = 23820344
api_hash = "df4339ef81253bad2463a65ae5b7b300"
bot_token = "7394299007:AAFft8frnlrKX_tUGMknwoHdSLoxWKRQvWc"

app = Client("bot_peliculas_lista", api_id=api_id, api_hash=api_hash, bot_token=bot_token)

excel_file = "canales_creados.xlsx"

user_results = {}
user_pages = {}

# ---------- NORMALIZAR TEXTO ----------
def normalizar(texto):
    texto = str(texto).lower()
    texto = unicodedata.normalize('NFD', texto)
    return ''.join(c for c in texto if unicodedata.category(c) != 'Mn')


# ---------- CARGAR EXCEL ----------
def cargar_peliculas():

    if not os.path.exists(excel_file):
        return []

    wb = load_workbook(excel_file)
    ws = wb.active

    peliculas = []

    for row in ws.iter_rows(min_row=2, values_only=True):

        nombre, enlace, genero = row[:3]

        if not nombre or not enlace:
            continue

        peliculas.append({
            "nombre": str(nombre),
            "enlace": str(enlace),
            "genero": str(genero) if genero else ""
        })

    return peliculas


# ---------- BOTONES PROMOCION ----------
def botones_promocion():

    botones = [

        [InlineKeyboardButton("🎬 Página Películas Melgar",
        url="https://tecnologysmith.github.io/Peliculas_Melgar.html")],

        [InlineKeyboardButton("🚀 TecnologySmith",
        url="https://tecnologysmith.godaddysites.com/")],

        [InlineKeyboardButton("🍿 Compra Netflix, Disney, HBO y más",
        url="https://tecnologysmith.github.io/plataformas.html/")],

        [InlineKeyboardButton("📥 Descárgalas en Terabox",
        url="https://1024terabox.com/s/14McCw4X4gtwraY07xGHJ5Q")],

        [InlineKeyboardButton("🎁 Sorteos en WhatsApp",
        url="https://chat.whatsapp.com/GYt7JHkJanz3P4tTkG3T5e")]

    ]

    return botones


# ---------- MENU ----------
def menu_principal():

    botones = [
        [InlineKeyboardButton("🎲 Película Aleatoria", callback_data="aleatoria")]
    ]

    return InlineKeyboardMarkup(botones)


# ---------- START ----------
@app.on_message(filters.command("start"))
def start(client, message):

    message.reply(
        "🍿 Bienvenido al buscador de películas\n\n"
        "🔎 Escribe el nombre de una película o género",
        reply_markup=menu_principal()
    )


# ---------- MOSTRAR LISTA ----------
def mostrar_lista(client, chat_id, user_id):

    resultados = user_results.get(user_id, [])

    if not resultados:
        client.send_message(chat_id, "❌ No hay resultados.")
        return

    pagina = user_pages.get(user_id, 0)

    por_pagina = 10

    inicio = pagina * por_pagina
    fin = inicio + por_pagina

    lote = resultados[inicio:fin]

    texto = ""

    texto += (
        "🎬 <b>Con @Mr_smithht en Películas Melgar</b>\n"
        "Encontramos estas películas que te pueden gustar:\n\n"
    )

    for i, peli in enumerate(lote, start=inicio + 1):
        texto += f'{i}. <a href="{peli["enlace"]}">{peli["nombre"]}</a>\n'

    total_paginas = (len(resultados) + por_pagina - 1) // por_pagina
    pagina_actual = pagina + 1

    texto += f"\n📄 Página {pagina_actual}/{total_paginas}\n\n"

    texto += (
        "🔎 También puedes buscar por género: acción, terror, comedia, etc.\n\n"
        "❓ Si no encontraste la película que deseas puedes escribir al creador del grupo:\n"
        "@Mr_smithht"
    )

    botones = []

    if fin < len(resultados):
        botones.append([InlineKeyboardButton("➡ Siguiente", callback_data="siguiente")])

    botones.append([InlineKeyboardButton("🎲 Aleatoria", callback_data="aleatoria")])

    teclado = botones + botones_promocion()

    client.send_message(
        chat_id,
        texto,
        parse_mode=ParseMode.HTML,
        disable_web_page_preview=True,
        reply_markup=InlineKeyboardMarkup(teclado)
    )


# ---------- MOSTRAR SUGERENCIAS ----------
def mostrar_sugerencias(client, chat_id):

    peliculas = cargar_peliculas()

    if not peliculas:
        client.send_message(chat_id, "No hay películas disponibles.")
        return

    sugerencias = random.sample(peliculas, min(10, len(peliculas)))

    texto = (
        "❌ <b>No encontramos tu película</b>\n\n"
        "Pero encontramos estas que te pueden gustar:\n\n"
    )

    for i, peli in enumerate(sugerencias, start=1):
        texto += f'{i}. <a href="{peli["enlace"]}">{peli["nombre"]}</a>\n'

    texto += (
        "\n\n🔎 También puedes buscar por género: acción, terror, comedia, etc.\n\n"
        "❓ Si no encontraste la película que deseas puedes escribir al creador del grupo:\n"
        "@Mr_smithht"
    )

    botones = [
        [InlineKeyboardButton("🎲 Otra Aleatoria", callback_data="aleatoria")]
    ]

    teclado = botones + botones_promocion()

    client.send_message(
        chat_id,
        texto,
        parse_mode=ParseMode.HTML,
        disable_web_page_preview=True,
        reply_markup=InlineKeyboardMarkup(teclado)
    )


# ---------- BUSCADOR ----------
@app.on_message(filters.text & (filters.private | filters.group))
def buscador(client, message):

    texto = normalizar(message.text)

    if len(texto) < 3:
        return

    peliculas = cargar_peliculas()

    palabras = texto.split()

    resultados = []

    for peli in peliculas:

        contenido = normalizar(
            f"{peli['nombre']} {peli['genero']}"
        )

        if all(p in contenido for p in palabras):
            resultados.append(peli)

    user_id = message.from_user.id if message.from_user else message.chat.id

    if not resultados:
        mostrar_sugerencias(client, message.chat.id)
        return

    random.shuffle(resultados)

    user_results[user_id] = resultados
    user_pages[user_id] = 0

    mostrar_lista(client, message.chat.id, user_id)


# ---------- SIGUIENTE PAGINA ----------
@app.on_callback_query(filters.regex("^siguiente$"))
def siguiente(client, callback_query):

    user_id = callback_query.from_user.id if callback_query.from_user else callback_query.message.chat.id

    user_pages[user_id] += 1

    mostrar_lista(client, callback_query.message.chat.id, user_id)

    callback_query.answer()


# ---------- PELICULA ALEATORIA ----------
@app.on_callback_query(filters.regex("^aleatoria$"))
def aleatoria(client, callback_query):

    peliculas = cargar_peliculas()

    if not peliculas:
        callback_query.answer("No hay películas.")
        return

    peli = random.choice(peliculas)

    texto = f'🎬 <a href="{peli["enlace"]}">{peli["nombre"]}</a>'

    client.send_message(
        callback_query.message.chat.id,
        texto,
        parse_mode=ParseMode.HTML,
        disable_web_page_preview=True
    )

    callback_query.answer()


print("🎬 Bot de películas iniciado correctamente")

app.run()